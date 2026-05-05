import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime
from collections import defaultdict

st.set_page_config(
    page_title="Energy Weekly Report Generator",
    page_icon="⚡",
    layout="wide"
)

st.markdown("""
<style>
.main-title  { font-size:2rem; font-weight:700; color:#1e3a5f; margin-bottom:.2rem; }
.subtitle    { font-size:1rem; color:#666; margin-bottom:1.5rem; }
.step-box    { background:#f0f4ff; border-left:4px solid #2563eb;
               border-radius:6px; padding:.8rem 1rem; margin-bottom:.8rem; }
.success-box { background:#f0fdf4; border-left:4px solid #16a34a;
               border-radius:6px; padding:.8rem 1rem; }
.warn-box    { background:#fffbeb; border-left:4px solid #d97706;
               border-radius:6px; padding:.8rem 1rem; margin-bottom:.6rem; }
</style>
""", unsafe_allow_html=True)

THAI_HOLIDAYS_2026 = {
    datetime.date(2026,  1,  1): "วันขึ้นปีใหม่",
    datetime.date(2026,  3,  4): "วันมาฆบูชา",
    datetime.date(2026,  4,  6): "วันจักรี",
    datetime.date(2026,  4, 13): "วันสงกรานต์",
    datetime.date(2026,  4, 14): "วันสงกรานต์",
    datetime.date(2026,  4, 15): "วันสงกรานต์",
    datetime.date(2026,  5,  1): "วันแรงงาน",
    datetime.date(2026,  5,  4): "วันฉัตรมงคล",
    datetime.date(2026,  5, 11): "วันวิสาขบูชา",
    datetime.date(2026,  6,  3): "วันเฉลิมพระชนมพรรษา ร.10",
    datetime.date(2026,  7, 10): "วันอาสาฬหบูชา",
    datetime.date(2026,  7, 11): "วันเข้าพรรษา",
    datetime.date(2026,  7, 28): "วันเฉลิมพระชนมพรรษา (ชดเชย)",
    datetime.date(2026,  8, 12): "วันแม่แห่งชาติ",
    datetime.date(2026, 10, 13): "วันคล้ายวันสวรรคต ร.9",
    datetime.date(2026, 10, 23): "วันปิยมหาราช",
    datetime.date(2026, 12,  5): "วันพ่อแห่งชาติ",
    datetime.date(2026, 12, 10): "วันรัฐธรรมนูญ",
    datetime.date(2026, 12, 31): "วันสิ้นปี",
}

DAY_TH = {0:"จ", 1:"อ", 2:"พ", 3:"พฤ", 4:"ศ", 5:"ส", 6:"อา"}

# ── Meter Mapping (Main Group / Group) ────────────────────────────────────────
METER_MAPPING = {
    "Meter_5":             {"main_group": "Extruder",              "group": "Extruder Line 3"},
    "Meter_6":             {"main_group": "Extruder",              "group": "Extruder Line 4"},
    "Meter_7":             {"main_group": "Extruder",              "group": "Extruder Line 3"},
    "Meter_8":             {"main_group": "Extruder",              "group": "Extruder Line 4"},
    "Meter_9":             {"main_group": "Extruder",              "group": "Extruder Line 2"},
    "Meter_12":            {"main_group": "Extruder",              "group": "Extruder Line 1"},
    "Meter_13":            {"main_group": "Extruder",              "group": "Extruder Line 1"},
    "Meter_24":            {"main_group": "Packing",               "group": "Packing"},
    "MCC3_1":              {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MCC3_2":              {"main_group": "Extruder",              "group": "Extruder Line 7"},
    "STOLZ#1_EX7":         {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "STOLZ#2_EX5,8":       {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Motor_Ext7":          {"main_group": "Extruder",              "group": "Extruder Line 7"},
    "MCC4_1":              {"main_group": "Receiving Raw Material", "group": "Receiving Raw Material"},
    "MCC4_2":              {"main_group": "Pre-Grinding",          "group": "Pre-Grinding"},
    "MCC4_3":              {"main_group": "Pre-Grinding",          "group": "Pre-Grinding"},
    "MCC4_4":              {"main_group": "Pre-Grinding",          "group": "Pre-Grinding"},
    "MCC5_1":              {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "MCC5_2":              {"main_group": "Bulk Tower",            "group": "Bulk Tower"},
    "MCC5_3":              {"main_group": "Bulk Tower",            "group": "Bulk Tower"},
    "MCC5_4":              {"main_group": "Receiving Raw Material", "group": "Receiving Raw Material"},
    "MCC5_5":              {"main_group": "Bulk Tower",            "group": "Bulk Tower"},
    "MCC5_6":              {"main_group": "Air Compressor",        "group": "Utility AC 5-6"},
    "Motor_Ext8":          {"main_group": "Extruder",              "group": "Extruder Line 8"},
    "MCC8_1":              {"main_group": "Extruder",              "group": "Extruder Line 8"},
    "DB_Ext5":             {"main_group": "Extruder",              "group": "Extruder Line 5"},
    "MCC6_2":              {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "DB_Ext9":             {"main_group": "Extruder",              "group": "Extruder Line 9"},
    "AirComp_P7":          {"main_group": "Air Compressor",        "group": "Utility AC 7"},
    "AirComp_P8":          {"main_group": "Air Compressor",        "group": "Utility AC 8"},
    "AirComp_P1234":       {"main_group": "Air Compressor",        "group": "Utility AC 1234"},
    "DB_24BIN":            {"main_group": "Packing",               "group": "Packing"},
    "MCC_40BIN":           {"main_group": "Packing",               "group": "Packing"},
    "Meter1_Grind3,4":     {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter_2_Grind3":      {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter_3_Grind4":      {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter_4_IDAH17":      {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter_30_WH2":        {"main_group": "FG WH",                 "group": "FG WH"},
    "Meter_36_AP6":        {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter_37_GD9_Sy":     {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "Meter_10_Repack":     {"main_group": "Packing",               "group": "Packing"},
    "Meter28_Grind11":     {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter26_Grind12":     {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT__15_LT_Feed":      {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "MT_16_Intake1_2":     {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "MT_17_GD_intake":     {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "MT_18_Mixer":         {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "MT_19_GDSys6_10":     {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_20_Grind_6":       {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_21_Grind_10":      {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_22_Coolroom":      {"main_group": "Coolroom",              "group": "Coolroom"},
    "MT_25GDSys11_12":     {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_26_Grind12":       {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_27_GD_AP6":        {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_28_Grind11":       {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "MT_29_WH1":           {"main_group": "Packing",               "group": "Packing"},
    "MT_32_Office":        {"main_group": "Office",                "group": "Office"},
    "MT_33_ENG":           {"main_group": "Maintenance",           "group": "Maintenance"},
    "MT_34_LT_WH3":        {"main_group": "Receiving Raw Material", "group": "Receiving Raw Material"},
    "MT_38_Farm":          {"main_group": "Farm",                  "group": "Farm"},
    "Meter_37":            {"main_group": "Batching / Pet-Batch",  "group": "Batching / Pet-Batch"},
    "Meter_GD9":           {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Stolz_609":           {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "Meter11_New_GD":      {"main_group": "Fine grinding",         "group": "Fine grinding"},
    "GD_coolroom":         {"main_group": "Coolroom",              "group": "Coolroom"},
    "New_24_bin":          {"main_group": "Packing",               "group": "Packing"},
    "MCC_PelletMill2":     {"main_group": "Bulk Tower",            "group": "Bulk Tower"},
    "Calculate Ex6":       {"main_group": "Extruder",              "group": "Extruder Line 6"},
    # ── NEW ──────────────────────────────────────────────────────────────────
    "Calculate Ex5":       {"main_group": "Extruder",              "group": "Extruder Line 5"},
    # ─────────────────────────────────────────────────────────────────────────
    "Meter_24 ลบ AC78":   {"main_group": "Packing",               "group": "Packing"},
}

# ── Allowed meter names — fixed order for output ──────────────────────────────
# NOTE: DB_Ext5, Meter_24, and DB_24BIN are intentionally EXCLUDED from this
#       list — they are used as source meters for virtual calculations only.
ALLOWED_METERS = [
    "Meter_5",
    "Meter_6",
    "Meter_7",
    "Meter_8",
    "Meter_9",
    "Meter_12",
    "Meter_13",
    # "Meter_24"    ← removed (source for Calculate Ex5 & Meter_24 ลบ AC78)
    "MCC3_1",
    "MCC3_2",
    "STOLZ#1_EX7",
    "STOLZ#2_EX5,8",
    "Motor_Ext7",
    "MCC4_1",
    "MCC4_2",
    "MCC4_3",
    "MCC4_4",
    "MCC5_1",
    "MCC5_2",
    "MCC5_3",
    "MCC5_4",
    "MCC5_5",
    "MCC5_6",
    "Motor_Ext8",
    "MCC8_1",
    # "DB_Ext5"     ← removed (source for Calculate Ex5)
    "MCC6_2",
    "DB_Ext9",
    "AirComp_P7",
    "AirComp_P8",
    "AirComp_P1234",
    # "DB_24BIN"    ← removed (source for Calculate Ex5)
    "MCC_40BIN",
    "Meter1_Grind3,4",
    "Meter_2_Grind3",
    "Meter_3_Grind4",
    "Meter_4_IDAH17",
    "Meter_30_WH2",
    "Meter_36_AP6",
    "Meter_37_GD9_Sy",
    "Meter_10_Repack",
    "Meter28_Grind11",
    "Meter26_Grind12",
    "MT__15_LT_Feed",
    "MT_16_Intake1_2",
    "MT_17_GD_intake",
    "MT_18_Mixer",
    "MT_19_GDSys6_10",
    "MT_20_Grind_6",
    "MT_21_Grind_10",
    "MT_22_Coolroom",
    "MT_25GDSys11_12",
    "MT_26_Grind12",
    "MT_27_GD_AP6",
    "MT_28_Grind11",
    "MT_29_WH1",
    "MT_32_Office",
    "MT_33_ENG",
    "MT_34_LT_WH3",
    "MT_38_Farm",
    "Meter_37",
    "Meter_GD9",
    "Stolz_609",
    "Meter11_New_GD",
    "GD_coolroom",
    "New_24_bin",
    "MCC_PelletMill2",
    "Calculate Ex6",
    "Calculate Ex5",          # ← NEW virtual meter
    "Meter_24 ลบ AC78",
]

# ── Virtual meter definitions ─────────────────────────────────────────────────
VIRTUAL_METERS = {
    "MDB6 Adjust": {
        "formula":       "MDB6 * 1.5",
        "base":          "MDB6",
        "factor":        1.5,
        "subtract_list": [],
    },
    "Calculate Ex6": {
        "formula":       "MDB6 Adjust - MCC6_2",
        "base":          "MDB6 Adjust",
        "subtract_list": ["MCC6_2"],
    },
    # ── NEW ──────────────────────────────────────────────────────────────────
    "Calculate Ex5": {
        "formula":       "DB_Ext5 - DB_24BIN",
        "base":          "DB_Ext5",
        "subtract_list": ["DB_24BIN"],
    },
    # ─────────────────────────────────────────────────────────────────────────
    "Meter_24 ลบ AC78": {
        "formula":       "Meter_24 - AirComp_P7 - AirComp_P8",
        "base":          "Meter_24",
        "subtract_list": ["AirComp_P7", "AirComp_P8"],
    },
}


def is_on_peak(ts, on_start: int, on_end: int, holidays: set) -> bool:
    if ts is None:
        return False
    d = ts.date()
    if d in holidays:
        return False
    if ts.weekday() >= 5:
        return False
    h = ts.hour + ts.minute / 60.0
    return on_start <= h < on_end


def interpolate_missing(kw_values: list, method: str) -> list:
    n = len(kw_values)
    arr = []
    for v in kw_values:
        if v is None or (isinstance(v, (int, float)) and v < 0):
            arr.append(float("nan"))
        else:
            arr.append(float(v))

    if method == "zero":
        return [0.0 if np.isnan(x) else x for x in arr]

    if method == "forward":
        last = 0.0
        out = []
        for x in arr:
            if np.isnan(x):
                out.append(last)
            else:
                last = x
                out.append(x)
        return out

    for i in range(n):
        if not np.isnan(arr[i]):
            continue
        prev_i = prev_v = next_i = next_v = None
        for j in range(i - 1, -1, -1):
            if not np.isnan(arr[j]):
                prev_i, prev_v = j, arr[j]
                break
        for j in range(i + 1, n):
            if not np.isnan(arr[j]):
                next_i, next_v = j, arr[j]
                break
        if prev_v is not None and next_v is not None:
            ratio = (i - prev_i) / (next_i - prev_i)
            arr[i] = prev_v + ratio * (next_v - prev_v)
        elif prev_v is not None:
            arr[i] = prev_v
        elif next_v is not None:
            arr[i] = next_v
        else:
            arr[i] = 0.0
    return arr


def count_bad(kw_values: list) -> int:
    return sum(1 for v in kw_values
               if v is None or (isinstance(v, (int, float)) and v < 0))


def get_week_ranges(timestamps: list) -> list:
    valid = sorted({ts.date() for ts in timestamps if ts is not None})
    if not valid:
        return []
    week_start = valid[0]
    last = valid[-1]
    weeks = []
    while week_start <= last:
        week_end = week_start + datetime.timedelta(days=6)
        weeks.append((week_start, week_end))
        week_start += datetime.timedelta(days=7)
    return weeks


def get_day_list(timestamps: list) -> list:
    valid = sorted({ts.date() for ts in timestamps if ts is not None})
    return valid


def compute_virtual_meters(results: dict, weeks: list, day_list: list):
    for vname, vdef in VIRTUAL_METERS.items():
        base_name = vdef["base"]
        sub_list  = vdef.get("subtract_list", [])

        if base_name not in results:
            continue

        base   = results[base_name]
        factor = vdef.get("factor", 1.0)

        w_on  = [round(v * factor, 4) for v in base["week_on"]]
        w_off = [round(v * factor, 4) for v in base["week_off"]]
        d_on  = [round(v * factor, 4) for v in base["day_on"]]
        d_off = [round(v * factor, 4) for v in base["day_off"]]

        for sub_name in sub_list:
            if sub_name in results:
                sub   = results[sub_name]
                w_on  = [max(0.0, a - b) for a, b in zip(w_on,  sub["week_on"])]
                w_off = [max(0.0, a - b) for a, b in zip(w_off, sub["week_off"])]
                d_on  = [max(0.0, a - b) for a, b in zip(d_on,  sub["day_on"])]
                d_off = [max(0.0, a - b) for a, b in zip(d_off, sub["day_off"])]

        results[vname] = {
            "no":       "",
            "week_on":  w_on,
            "week_off": w_off,
            "day_on":   d_on,
            "day_off":  d_off,
            "n_miss":   0,
            "virtual":  True,
            "formula":  vdef["formula"],
        }

    return results


def process_file(uploaded_file, on_start, on_end, holidays, fill_method):
    wb = load_workbook(uploaded_file, data_only=True)
    if "RawData" not in wb.sheetnames:
        return None, "ไม่พบ sheet 'RawData'"

    ws = wb["RawData"]
    ts_row = list(ws.iter_rows(min_row=2, max_row=2, min_col=4, values_only=True))[0]
    timestamps = list(ts_row)

    valid_ts = [t for t in timestamps if t is not None]
    if not valid_ts:
        return None, "ไม่พบ timestamps"

    weeks    = get_week_ranges(valid_ts)
    day_list = get_day_list(valid_ts)

    if not weeks:
        return None, "ไม่สามารถกำหนดสัปดาห์ได้"

    meter_order = []
    seen = set()
    kw_map = {}
    no_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, dtype, no = row[0], row[1], row[2]
        if name is None or dtype != "Kw":
            continue
        if name not in seen:
            seen.add(name)
            meter_order.append(name)
        kw_map[name] = list(row[3:])
        no_map[name] = no

    if not kw_map:
        return None, "ไม่พบข้อมูล Kw"

    results = {}
    total_missing = 0
    n_weeks  = len(weeks)
    n_days   = len(day_list)
    day_idx  = {d: i for i, d in enumerate(day_list)}

    for name in meter_order:
        raw   = kw_map[name]
        n_bad = count_bad(raw)
        total_missing += n_bad
        clean = interpolate_missing(raw, fill_method)

        w_on  = [0.0] * n_weeks
        w_off = [0.0] * n_weeks
        d_on  = [0.0] * n_days
        d_off = [0.0] * n_days

        for i, ts in enumerate(timestamps):
            if ts is None or i >= len(clean):
                continue
            kw = clean[i]
            if np.isnan(kw) or kw < 0:
                continue
            energy  = kw * 0.25
            on      = is_on_peak(ts, on_start, on_end, holidays)
            ts_date = ts.date()

            for wi, (ws2, we) in enumerate(weeks):
                if ws2 <= ts_date <= we:
                    (w_on if on else w_off)[wi] += energy
                    break

            if ts_date in day_idx:
                di = day_idx[ts_date]
                (d_on if on else d_off)[di] += energy

        results[name] = {
            "no":       no_map[name],
            "week_on":  w_on,
            "week_off": w_off,
            "day_on":   d_on,
            "day_off":  d_off,
            "n_miss":   n_bad,
        }

    # ── Append virtual meters ──────────────────────────────────────────────
    results = compute_virtual_meters(results, weeks, day_list)

    # ── Append virtual meter names to meter_order ──────────────────────────
    for vname in VIRTUAL_METERS:
        if vname in results and vname not in meter_order:
            meter_order.append(vname)

    return {
        "meter_order":   meter_order,
        "results":       results,
        "weeks":         weeks,
        "day_list":      day_list,
        "total_missing": total_missing,
    }, None


def merge_all(all_data: list) -> dict:
    if len(all_data) == 1:
        return all_data[0]

    all_weeks_set = set()
    all_days_set  = set()
    for d in all_data:
        all_weeks_set.update(d["weeks"])
        all_days_set.update(d["day_list"])
    all_weeks = sorted(all_weeks_set, key=lambda x: x[0])
    all_days  = sorted(all_days_set)
    n_weeks   = len(all_weeks)
    n_days    = len(all_days)
    day_idx   = {d: i for i, d in enumerate(all_days)}

    seen = set()
    all_meters = []
    for d in all_data:
        for name in d["meter_order"]:
            if name not in seen:
                seen.add(name)
                all_meters.append(name)

    merged = {}
    for name in all_meters:
        w_on  = [0.0] * n_weeks
        w_off = [0.0] * n_weeks
        d_on  = [0.0] * n_days
        d_off = [0.0] * n_days
        n_miss = 0
        no = None

        for d in all_data:
            if name not in d["results"]:
                continue
            r  = d["results"][name]
            no = r["no"]
            n_miss += r["n_miss"]

            for local_wi, wk in enumerate(d["weeks"]):
                try:
                    global_wi = all_weeks.index(wk)
                    w_on[global_wi]  += r["week_on"][local_wi]
                    w_off[global_wi] += r["week_off"][local_wi]
                except ValueError:
                    pass

            for local_di, day in enumerate(d["day_list"]):
                if day in day_idx:
                    global_di = day_idx[day]
                    d_on[global_di]  += r["day_on"][local_di]
                    d_off[global_di] += r["day_off"][local_di]

        merged[name] = {
            "no":       no,
            "week_on":  w_on,
            "week_off": w_off,
            "day_on":   d_on,
            "day_off":  d_off,
            "n_miss":   n_miss,
        }

    return {
        "meter_order":   all_meters,
        "results":       merged,
        "weeks":         all_weeks,
        "day_list":      all_days,
        "total_missing": sum(d["total_missing"] for d in all_data),
    }


def get_filtered_meter_order(data: dict) -> list:
    """
    Return meters in ALLOWED_METERS order, keeping only those present in results.
    Meters not in ALLOWED_METERS are excluded entirely
    (includes DB_Ext5, Meter_24, DB_24BIN which are source-only meters).
    """
    results = data["results"]
    return [name for name in ALLOWED_METERS if name in results]


def get_meter_info(name: str):
    info = METER_MAPPING.get(name, {})
    return info.get("main_group", ""), info.get("group", "")


def is_virtual(name: str, results: dict = None) -> bool:
    if results is None:
        return False
    return results.get(name, {}).get("virtual", False)


def build_excel(data: dict, month_year: str) -> io.BytesIO:
    weeks    = data["weeks"]
    day_list = data["day_list"]
    results  = data["results"]

    # ── Use filtered + ordered meter list ─────────────────────────────────
    meter_order = get_filtered_meter_order(data)

    n_days = len(day_list)

    DARK         = "1e3a5f"
    FILLS        = ["dbeafe", "d1fae5", "fef3c7", "ede9fe", "fee2e2",
                    "fce7f3", "ecfdf5", "fff7ed", "f0fdf4", "e0f2fe"]
    ALT          = "f8fafc"
    OFF_DAY_FILL = "e2e8f0"
    VIRTUAL_FILL = "fff9c4"

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    df = Font(name="Arial", size=9)
    bf = Font(name="Arial", size=9, bold=True)
    vf = Font(name="Arial", size=9, italic=True, color="555555")

    def fi(h):
        return PatternFill("solid", fgColor=h)

    def bo(s="thin"):
        t = Side(style=s)
        return Border(left=t, right=t, top=t, bottom=t)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet: Daily ─────────────────────────────────────────────────────────
    DATA_START_COL = 5
    ws3 = wb.create_sheet("Daily")

    # Row 1: date group headers (3 cols per day)
    col = DATA_START_COL
    for di, day in enumerate(day_list):
        ws3.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        day_name = DAY_TH[day.weekday()]
        label    = f"{day.strftime('%d/%m')}\n({day_name})"
        c = ws3.cell(1, col, label)
        wi_color = next(
            (wi for wi, (ws_d, we_d) in enumerate(weeks) if ws_d <= day <= we_d), 0
        )
        is_weekend = day.weekday() >= 5
        c.fill = fi(OFF_DAY_FILL) if is_weekend else fi(FILLS[wi_color % len(FILLS)])
        c.font = Font(
            name="Arial", bold=True, size=9,
            color="888888" if is_weekend else DARK
        )
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col += 3

    # Grand Total header
    ws3.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    c = ws3.cell(1, col, f"Grand Total ({month_year})")
    c.fill = fi(DARK)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: column headers
    sh3 = ["Meter Name", "No.", "Main Group", "Group"]
    for _ in range(n_days):
        sh3 += ["On Peak\n(kWh)", "Off Peak\n(kWh)", "Total\n(kWh)"]
    sh3 += ["On Peak\n(kWh)", "Off Peak\n(kWh)", "Total\n(kWh)"]
    for ci, h in enumerate(sh3, 1):
        c = ws3.cell(2, ci, h)
        c.fill = fi(DARK)
        c.font = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bo()

    # Data rows
    for ri, name in enumerate(meter_order):
        if name not in results:
            continue
        r        = results[name]
        rn       = 3 + ri
        main_grp, grp = get_meter_info(name)
        virtual  = is_virtual(name, results)
        alt      = fi(VIRTUAL_FILL) if virtual else (fi(ALT) if ri % 2 == 0 else PatternFill())

        def dc(ci, val, fmt=None, bold=False, bg=None):
            c = ws3.cell(rn, ci, val)
            c.border = bo()
            c.font   = bf if bold else (vf if virtual else df)
            c.alignment = Alignment(
                horizontal="right" if isinstance(val, float) else "left"
            )
            if fmt:
                c.number_format = fmt
            if bg:
                c.fill = fi(bg)
            else:
                c.fill = alt

        dc(1, name);      ws3.cell(rn, 1).fill = alt
        dc(2, r["no"]);   ws3.cell(rn, 2).fill = alt
        dc(3, main_grp);  ws3.cell(rn, 3).fill = alt
        dc(4, grp);       ws3.cell(rn, 4).fill = alt

        ton = toff = 0.0
        col = DATA_START_COL
        for di, day in enumerate(day_list):
            ov = round(r["day_on"][di], 1)  if di < len(r["day_on"])  else 0.0
            fv = round(r["day_off"][di], 1) if di < len(r["day_off"]) else 0.0
            tv = round(ov + fv, 1)
            ton  += ov
            toff += fv
            is_weekend = day.weekday() >= 5
            bg_day = OFF_DAY_FILL if is_weekend else None
            for v in [ov, fv, tv]:
                c = ws3.cell(rn, col, v)
                c.border = bo()
                c.font   = vf if virtual else df
                c.number_format = "#,##0.0"
                c.alignment = Alignment(horizontal="right")
                c.fill = fi(bg_day) if bg_day else alt
                col += 1

        for v in [round(ton, 1), round(toff, 1), round(ton + toff, 1)]:
            c = ws3.cell(rn, col, v)
            c.border = bo()
            c.font   = bf
            c.number_format = "#,##0.0"
            c.alignment = Alignment(horizontal="right")
            c.fill = fi("e8f0fe")
            col += 1

    # Column widths
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 5
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 22
    for ci in range(DATA_START_COL, DATA_START_COL + (n_days + 1) * 3):
        ws3.column_dimensions[get_column_letter(ci)].width = 10
    ws3.row_dimensions[1].height = 35
    ws3.row_dimensions[2].height = 40
    ws3.freeze_panes = "E3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── UI ───────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">⚡ Energy Daily Report Generator</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="subtitle">แปลงไฟล์ ReportRaw → Energy Daily Report '
    '(On Peak / Off Peak) · Main Group / Group · Calculated Meters</div>',
    unsafe_allow_html=True
)

with st.sidebar:
    st.header("⚙️ ตั้งค่า")
    month_year = st.text_input("ป้ายกำกับรายงาน", value="March 2026")

    st.markdown("---")
    st.subheader("🕐 ช่วง On Peak (วันธรรมดา)")
    c1, c2 = st.columns(2)
    with c1:
        on_start = st.number_input("เริ่ม (ชม.)", 0, 23, value=9)
    with c2:
        on_end = st.number_input("สิ้นสุด (ชม.)", 1, 24, value=22)
    st.caption(f"On Peak: **{on_start:02d}:00 – {on_end:02d}:00** จ–ศ")

    st.markdown("---")
    st.subheader("📅 วันหยุดนักขัตฤกษ์")
    use_thai = st.checkbox("ใช้วันหยุดไทย 2026 (MEA/PEA)", value=False)
    extra_text = st.text_area(
        "เพิ่มวันหยุดพิเศษ (YYYY-MM-DD แต่ละบรรทัด)",
        placeholder="2026-03-04\n2026-04-06",
        height=90,
    )
    holidays: set = set()
    if use_thai:
        holidays.update(THAI_HOLIDAYS_2026.keys())
    parse_errors = []
    for line in extra_text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            holidays.add(datetime.date.fromisoformat(line))
        except ValueError:
            parse_errors.append(line)
    if parse_errors:
        st.warning(f"รูปแบบวันที่ไม่ถูกต้อง: {', '.join(parse_errors)}")
    if holidays:
        st.success(f"✅ วันหยุด: {len(holidays)} วัน")
        with st.expander("ดูรายการวันหยุด"):
            for d in sorted(holidays):
                label = THAI_HOLIDAYS_2026.get(d, "วันหยุดพิเศษ")
                st.write(f"• {d.strftime('%d %b %Y')} — {label}")
    else:
        st.info("ไม่มีวันหยุดพิเศษ (ส–อ = Off Peak เท่านั้น)")

    st.markdown("---")
    st.subheader("🔧 Slot ที่ขาดหาย (ค่า -1)")
    fill_choice = st.radio(
        "วิธีจัดการ",
        options=[
            "เฉลี่ย slot ก่อน/หลัง (Linear)",
            "ใส่ค่า 0",
            "ใช้ค่า slot ก่อนหน้า (Forward fill)",
        ],
        index=0,
    )
    METHOD_MAP = {
        "เฉลี่ย slot ก่อน/หลัง (Linear)":     "linear",
        "ใส่ค่า 0":                             "zero",
        "ใช้ค่า slot ก่อนหน้า (Forward fill)": "forward",
    }
    fill_method = METHOD_MAP[fill_choice]

    st.markdown("---")
    st.subheader("🧮 Virtual / Calculated Meters")
    with st.expander("ดูสูตรคำนวณ"):
        st.markdown("""
| Meter | สูตร |
|---|---|
| **MDB6 Adjust** | MDB6 × 1.5 |
| **Calculate Ex6** | MDB6 Adjust − MCC6_2 |
| **Calculate Ex5** | DB_Ext5 − DB_24BIN |
| **Meter_24 ลบ AC78** | Meter_24 − AirComp_P7 − AirComp_P8 |
        """)
        st.caption("แถวสีเหลืองอ่อนในรายงาน = Calculated meter")
        st.caption("⚠️ DB_Ext5 · Meter_24 · DB_24BIN = source meters (ไม่แสดงในรายงาน)")

# ── Main ──────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="step-box">📁 <b>Step 1: อัพโหลดไฟล์ ReportRaw</b> '
    '— รองรับหลายไฟล์พร้อมกัน</div>',
    unsafe_allow_html=True
)

uploaded_files = st.file_uploader(
    "เลือกไฟล์ ReportRaw (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.markdown(
        f'<div class="success-box">✅ อัพโหลดแล้ว <b>{len(uploaded_files)} ไฟล์</b></div>',
        unsafe_allow_html=True
    )
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("ไฟล์", len(uploaded_files))
    c2.metric("On Peak", f"{on_start:02d}:00–{on_end:02d}:00")
    c3.metric("วันหยุด", f"{len(holidays)} วัน")
    c4.metric("Missing slot", fill_choice.split("(")[0].strip()[:15])

    st.markdown(
        '<div class="step-box">⚙️ <b>Step 2: ประมวลผล</b></div>',
        unsafe_allow_html=True
    )

    if st.button("🚀 สร้าง Energy Daily Report", type="primary", use_container_width=True):
        progress  = st.progress(0)
        status    = st.empty()
        all_data  = []
        has_error = False

        for idx, uf in enumerate(uploaded_files):
            status.text(f"⏳ {uf.name} ({idx+1}/{len(uploaded_files)})")
            progress.progress(idx / len(uploaded_files))
            data, err = process_file(uf, on_start, on_end, holidays, fill_method)
            if err:
                st.error(f"❌ {uf.name}: {err}")
                has_error = True
            else:
                n_m = data["total_missing"]
                filtered_count = len(get_filtered_meter_order(data))
                msg = (f"✅ {uf.name}: แสดง {filtered_count} meters, "
                       f"{len(data['weeks'])} สัปดาห์, {len(data['day_list'])} วัน")
                if n_m:
                    msg += f" ⚠️ {n_m} slots ขาดหาย (แก้ไขแล้ว)"
                st.success(msg)
                all_data.append(data)

        progress.progress(1.0)

        if all_data and not has_error:
            status.text("🔄 รวมข้อมูลและสร้าง Excel...")
            merged         = merge_all(all_data)
            filtered_order = get_filtered_meter_order(merged)
            buf            = build_excel(merged, month_year)
            status.empty()
            progress.empty()

            st.markdown("---")
            st.markdown(
                '<div class="step-box">📊 <b>Step 3: ผลลัพธ์</b></div>',
                unsafe_allow_html=True
            )

            total_miss = merged["total_missing"]
            if total_miss:
                st.markdown(
                    f'<div class="warn-box">⚠️ พบ <b>{total_miss} time slots</b> '
                    f'ที่ข้อมูลขาดหาย → แก้ไขด้วย: <b>{fill_choice}</b></div>',
                    unsafe_allow_html=True
                )

            grand_on  = sum(sum(merged["results"][n]["day_on"])  for n in filtered_order)
            grand_off = sum(sum(merged["results"][n]["day_off"]) for n in filtered_order)
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Meters (แสดง)",    len(filtered_order))
            c2.metric("วัน",              len(merged["day_list"]))
            c3.metric("Total On Peak",    f"{grand_on:,.0f} kWh")
            c4.metric("Total Off Peak",   f"{grand_off:,.0f} kWh")

            # Preview tab — Daily only
            st.subheader("📆 Daily Preview (15 meters แรก)")
            preview_d = []
            for name in filtered_order[:15]:
                r = merged["results"][name]
                info = METER_MAPPING.get(name, {})
                row = {
                    "Meter":      name,
                    "No.":        r["no"],
                    "Main Group": info.get("main_group", ""),
                    "Group":      info.get("group", ""),
                }
                for di, day in enumerate(merged["day_list"]):
                    on_v  = round(r["day_on"][di], 1)  if di < len(r["day_on"])  else 0.0
                    off_v = round(r["day_off"][di], 1) if di < len(r["day_off"]) else 0.0
                    label = day.strftime("%d/%m")
                    row[f"{label} On"]    = on_v
                    row[f"{label} Off"]   = off_v
                    row[f"{label} Total"] = round(on_v + off_v, 1)
                row["Grand Total"] = round(sum(r["day_on"]) + sum(r["day_off"]), 1)
                if r.get("virtual"):
                    row["📌 Formula"] = r.get("formula", "")
                preview_d.append(row)
            st.dataframe(pd.DataFrame(preview_d), use_container_width=True)

            # ── Department Usage Breakdown ────────────────────────────────
            st.subheader("🏭 Department Usage Breakdown")
            dept_totals = defaultdict(lambda: {"on": 0.0, "off": 0.0})
            for name in filtered_order:
                r = merged["results"][name]
                main_grp, _ = get_meter_info(name)
                if not main_grp:
                    main_grp = "Other"
                dept_totals[main_grp]["on"]  += sum(r["day_on"])
                dept_totals[main_grp]["off"] += sum(r["day_off"])

            dept_rows = []
            for dept, vals in dept_totals.items():
                total = vals["on"] + vals["off"]
                dept_rows.append({
                    "Main Group":       dept,
                    "On Peak (kWh)":    round(vals["on"],  1),
                    "Off Peak (kWh)":   round(vals["off"], 1),
                    "Total (kWh)":      round(total,       1),
                })
            dept_rows.sort(key=lambda x: x["Total (kWh)"], reverse=True)

            grand_total_all = sum(r["Total (kWh)"] for r in dept_rows)
            for r in dept_rows:
                r["% Share"] = f"{r['Total (kWh)'] / grand_total_all * 100:.1f}%" if grand_total_all else "0.0%"

            st.dataframe(
                pd.DataFrame(dept_rows),
                use_container_width=True,
                hide_index=True,
            )

            fname = f"Energy_Daily_{month_year.replace(' ', '_')}.xlsx"
            st.download_button(
                "⬇️ ดาวน์โหลด Energy Daily Report (.xlsx)",
                data=buf,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
            st.balloons()

        elif has_error:
            status.empty()
            progress.empty()
            st.error("กรุณาตรวจสอบไฟล์และลองใหม่")

else:
    st.info("👆 กรุณาอัพโหลดไฟล์ ReportRaw (.xlsx) เพื่อเริ่มต้น")
    with st.expander("📖 วิธีใช้งาน"):
        st.markdown("""
### โครงสร้างไฟล์ Input
- Sheet **`RawData`** มี:
  - Row 2 = timestamps ทุก 15 นาที (cols D เป็นต้นไป)
  - แต่ละ meter มี 3 แถว: `Kw` · `BeginKwhr` · `FinalKwhr`

### Output — 1 Sheet
| Sheet | เนื้อหา |
|---|---|
| `Daily` | On/Off Peak × **รายวัน** + Grand Total |

### Meter ที่แสดง (ตามลำดับที่กำหนด)
- แสดงเฉพาะ meters ใน ALLOWED_METERS เท่านั้น
- **DB_Ext5 · Meter_24 · DB_24BIN** = source meters (ไม่แสดงในรายงาน)

### Calculated Meters (สีเหลืองอ่อน)
| Meter | สูตร | Main Group | Group |
|---|---|---|---|
| Calculate Ex6 | MDB6 Adjust − MCC6_2 | Extruder | Extruder Line 6 |
| Calculate Ex5 | DB_Ext5 − DB_24BIN | Extruder | Extruder Line 5 |
| Meter_24 ลบ AC78 | Meter_24 − AirComp_P7 − AirComp_P8 | Packing | Packing |

### สีใน Sheet Daily
- **สีอ่อนตามสัปดาห์** = วันธรรมดา
- **สีเทา** = เสาร์/อาทิตย์
- **สีเหลืองอ่อน** = Calculated meter
        """)

st.markdown("---")
st.markdown(
    "<div style='text-align:center;color:#999;font-size:.8rem;'>"
    "⚡ Energy Daily Report Generator · On Peak 09:00–22:00 (MEA/PEA) · "
    "Daily Sheet · Main Group / Group · Calculated Meters"
    "</div>",
    unsafe_allow_html=True,
)
